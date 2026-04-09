#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
转换输出格式：合并 scorer_key_point 和 keypoint_results 字段
"""

import openpyxl
from typing import Dict, Any

def convert_result_format(old_row: Dict[str, Any]) -> Dict[str, Any]:
    """将旧格式转换为新格式"""
    new_row = dict(old_row)
    
    # 合并 scorer_key_point_1 和 scorer_key_point_0
    scorer_key_point_lines = []
    
    # 1分条件
    kp_1 = old_row.get('scorer_key_point_1', '')
    if kp_1:
        scorer_key_point_lines.append("【满足1分的条件】")
        items = [item.strip() for item in str(kp_1).split('|') if item.strip()]
        for i, item in enumerate(items, 1):
            scorer_key_point_lines.append(f"{i}. {item}")
    
    # 0分条件
    kp_0 = old_row.get('scorer_key_point_0', '')
    if kp_0:
        scorer_key_point_lines.append("\n【触发0分的条件】")
        items = [item.strip() for item in str(kp_0).split('|') if item.strip()]
        for i, item in enumerate(items, 1):
            scorer_key_point_lines.append(f"{i}. {item}")
    
    new_row['scorer_key_point'] = "\n".join(scorer_key_point_lines)
    
    # 合并 keypoint_1_results 和 keypoint_0_results
    keypoint_results_lines = []
    
    # 1分考点判定结果
    kpr_1 = old_row.get('keypoint_1_results', '')
    kpr_1_summary = old_row.get('keypoint_1_results_summary', '')
    if kpr_1:
        keypoint_results_lines.append("【1分考点判定结果】")
        # 解析原有格式
        sections = str(kpr_1).split('【1分考点】')
        idx = 1
        for section in sections[1:]:  # 跳过第一个空字符串
            lines = section.strip().split('\n')
            if len(lines) >= 3:
                point = lines[0].strip()
                satisfied = lines[1].replace('【是否满足】', '').strip()
                evidence = lines[2].replace('【判定依据】', '').strip()
                status = "✓ 满足" if str(satisfied).lower() in ("true", "1", "是", "满足") else "✗ 未满足"
                keypoint_results_lines.append(f"{idx}. {point}")
                keypoint_results_lines.append(f"   状态: {status}")
                keypoint_results_lines.append(f"   依据: {evidence}")
                keypoint_results_lines.append("")
                idx += 1
        new_row['keypoint_results_summary'] = f"1分考点: {kpr_1_summary}" if kpr_1_summary else ""
    
    # 0分考点判定结果
    kpr_0 = old_row.get('keypoint_0_results', '')
    if kpr_0:
        keypoint_results_lines.append("\n【0分考点判定结果】")
        # 解析原有格式
        sections = str(kpr_0).split('【0分考点】')
        idx = 1
        triggered_count = 0
        total_count = 0
        for section in sections[1:]:  # 跳过第一个空字符串
            lines = section.strip().split('\n')
            if len(lines) >= 3:
                point = lines[0].strip()
                satisfied = lines[1].replace('【是否满足】', '').strip()
                evidence = lines[2].replace('【判定依据】', '').strip()
                is_triggered = str(satisfied).lower() in ("true", "1", "是", "满足")
                if is_triggered:
                    triggered_count += 1
                total_count += 1
                status = "✓ 触发" if is_triggered else "✗ 未触发"
                keypoint_results_lines.append(f"{idx}. {point}")
                keypoint_results_lines.append(f"   状态: {status}")
                keypoint_results_lines.append(f"   依据: {evidence}")
                keypoint_results_lines.append("")
                idx += 1
        
        summary_0 = f"0分考点触发: {triggered_count}/{total_count}"
        if new_row.get('keypoint_results_summary'):
            new_row['keypoint_results_summary'] += f", {summary_0}"
        else:
            new_row['keypoint_results_summary'] = summary_0
    
    new_row['keypoint_results'] = "\n".join(keypoint_results_lines)
    
    # 删除旧字段
    for old_field in ['scorer_key_point_1', 'scorer_key_point_0', 
                      'keypoint_1_results', 'keypoint_1_results_summary', 'keypoint_0_results']:
        if old_field in new_row:
            del new_row[old_field]
    
    return new_row


def main():
    input_file = '/Users/zhujiangdi/Desktop/0401_豆包评估结果.xlsx'
    output_file = '/Users/zhujiangdi/Desktop/0401_豆包评估结果_新格式.xlsx'
    
    print(f"读取: {input_file}")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # 读取表头
    headers = [cell.value for cell in ws[1]]
    print(f"原列数: {len(headers)}")
    
    # 读取数据
    data = []
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for idx, cell in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = cell.value
        data.append(row_data)
    
    print(f"数据行数: {len(data)}")
    
    # 转换格式
    new_data = []
    for i, row in enumerate(data):
        new_row = convert_result_format(row)
        new_data.append(new_row)
        if (i + 1) % 10 == 0:
            print(f"已转换: {i + 1}/{len(data)}")
    
    # 保存新格式
    print(f"\n保存到: {output_file}")
    
    # 收集所有字段
    all_keys = set()
    for item in new_data:
        all_keys.update(item.keys())
    
    # 定义字段顺序
    priority_keys = [
        'session_id', 'query_id', 'query', 'answer',
        'kp_thinking', 'kp_analysis', 'kp_main_demand', 
        'kp_key_point_1', 'kp_key_point_0',
        'scorer_thinking', 'scorer_analysis', 'scorer_main_demand',
        'scorer_key_point',  # 合并后的字段
        'keypoint_results',  # 合并后的字段
        'keypoint_results_summary',
        'score', 'reason',
        'question_type', 'question_type_reason',
        'evaluated_at', 'error'
    ]
    
    new_headers = [k for k in priority_keys if k in all_keys]
    new_headers += [k for k in all_keys if k not in priority_keys]
    
    print(f"新列数: {len(new_headers)}")
    
    # 创建新工作簿
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # 写入表头
    new_ws.append(new_headers)
    
    # 写入数据
    for item in new_data:
        row = []
        for key in new_headers:
            value = item.get(key, "")
            if isinstance(value, (list, dict)):
                import json
                value = json.dumps(value, ensure_ascii=False)
            elif value is None:
                value = ""
            row.append(value)
        new_ws.append(row)
    
    new_wb.save(output_file)
    print(f"✅ 转换完成!")


if __name__ == '__main__':
    main()
