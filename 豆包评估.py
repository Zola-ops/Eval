#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
豆包API版本评估脚本 - 使用豆包API
使用方法:
    # 单条评估
    python standalone_evaluator_zhipu.py --query "问题" --answer "回答"

    # 批量评估
    python standalone_evaluator_zhipu.py --excel 考点摸底.xlsx --output results.xlsx

    # 自定义API Key
    python standalone_evaluator_zhipu.py --api_key "your-zhipu-key" --query "问题" --answer "回答"
"""

import os
import re
import sys
import json
import time
import argparse
from datetime import datetime
from typing import Dict, List, Any, Optional
import requests
import openpyxl
from jinja2 import Template


def clean_json_text(text: str) -> str:
    """清理JSON文本，移除markdown代码块标记并修复常见JSON问题"""
    if not text:
        return ""
    
    # 移除markdown代码块标记
    text = text.strip()
    if text.startswith('```json'):
        text = text[7:]
    if text.startswith('```'):
        text = text[3:]
    if text.endswith('```'):
        text = text[:-3]
    
    text = text.strip()
    
    # 尝试提取第一个完整的JSON对象
    start_idx = text.find('{')
    end_idx = text.rfind('}')
    
    if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
        text = text[start_idx:end_idx + 1]
    
    # 直接将所有换行符、回车符、制表符替换为空格（最简单有效的方法）
    # 这是因为智谱AI在JSON字符串值中插入了未转义的换行符
    text = text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    
    # 压缩多个连续空格为单个空格
    text = re.sub(r' +', ' ', text)
    
    return text.strip()


class APIClient:
    """豆包API客户端"""

    def __init__(self, api_url: str, api_token: str, model_config: Dict[str, Any]):
        self.api_url = api_url
        self.api_token = api_token
        self.model_config = model_config
        self.headers = {
            "Authorization": f"Bearer {api_token}",
            "Content-Type": "application/json"
        }

    def call_llm(self, system_prompt: str, user_prompt: str, 
                 context: Optional[str] = None) -> str:
        """
        调用豆包API - 标准OpenAI格式
        """
        # 构建messages
        messages = [
            {"role": "system", "content": system_prompt}
        ]
        
        # 如果有上下文，添加到system prompt
        if context:
            messages[0]["content"] = f"{system_prompt}\n\n{context}"
        
        messages.append({"role": "user", "content": user_prompt})
        
        # 构建请求payload
        payload = {
            "model": self.model_config.get("model", "glm-4.5"),
            "messages": messages,
            "temperature": self.model_config.get("temperature", 0.3),
            "top_p": self.model_config.get("top_p", 0.7),
            "max_tokens": self.model_config.get("max_completion_tokens", 2000),
            "stream": False
        }
        
        # 调用API
        try:
            response = requests.post(
                self.api_url,
                headers=self.headers,
                json=payload,
                timeout=120
            )
            response.raise_for_status()
            
            result = response.json()
            
            # 从响应中提取内容
            if "choices" in result and len(result["choices"]) > 0:
                content = result["choices"][0]["message"]["content"]
                return content
            else:
                raise ValueError(f"API响应格式异常: {result}")
                
        except requests.exceptions.RequestException as e:
            raise Exception(f"API调用失败: {str(e)}")
        except Exception as e:
            raise Exception(f"解析API响应失败: {str(e)}")


class KeypointGenerator:
    """考点生成器"""

    def __init__(self, api_client: APIClient, sp: str, up_template: str):
        self.api = api_client
        self.sp = sp
        self.up_template = up_template

    def _format_context(self, context: List[Dict[str, Any]]) -> str:
        """格式化上下文历史"""
        if not context:
            return ""
        
        context_parts = []
        for idx, item in enumerate(context, 1):
            context_parts.append(f"## 历史对话 {idx}")
            context_parts.append(f"问题: {item.get('query', '')}")
            context_parts.append(f"回答: {item.get('answer', '')}")
            if 'keypoints' in item:
                context_parts.append(f"关键点: {json.dumps(item['keypoints'], ensure_ascii=False)}")
        
        return "\n\n".join(context_parts)

    def generate(self, query: str, date: Optional[str] = None,
                demand: Optional[str] = None,
                context: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
        """生成考点"""
        # 使用Jinja2渲染用户提示词
        up_tpl = Template(self.up_template)
        up_content = up_tpl.render({
            "query": query,
            "date": date or "",
            "demand": demand or "",
            "context": self._format_context(context) if context else ""
        })
        
        # 调用API
        response_text = self.api.call_llm(self.sp, up_content)
        
        # 解析JSON
        try:
            cleaned = clean_json_text(response_text)
            return json.loads(cleaned)
        except json.JSONDecodeError as e:
            raise Exception(f"考点生成响应解析失败: {str(e)}")


class Scorer:
    """评分器"""

    def __init__(self, api_client: APIClient, sp: str, up_template: str):
        self.api = api_client
        self.sp = sp
        self.up_template = up_template

    def _format_context(self, context: List[Dict[str, Any]]) -> str:
        """格式化上下文历史"""
        if not context:
            return ""
        
        context_parts = []
        for idx, item in enumerate(context, 1):
            context_parts.append(f"## 历史对话 {idx}")
            context_parts.append(f"问题: {item.get('query', '')}")
            context_parts.append(f"回答: {item.get('answer', '')}")
            if 'keypoints' in item:
                context_parts.append(f"关键点: {json.dumps(item['keypoints'], ensure_ascii=False)}")
        
        return "\n\n".join(context_parts)

    def score(self, query: str, answer: str,
              keypoint_result: Dict[str, Any],
              context: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
        """评分"""
        # 获取具体考点列表
        kp_1_list = keypoint_result.get("key_point_1", [])
        kp_0_list = keypoint_result.get("key_point_0", [])
        
        # 渲染用户提示词
        up_tpl = Template(self.up_template)
        up_content = up_tpl.render({
            "query": query,
            "thinking": keypoint_result.get("thinking", ""),
            "analysis": keypoint_result.get("analysis", ""),
            "main_demand": keypoint_result.get("main_demand", ""),
            "key_point_1": kp_1_list if isinstance(kp_1_list, list) else [str(kp_1_list)],
            "key_point_0": kp_0_list if isinstance(kp_0_list, list) else [str(kp_0_list)],
            "key_point_1_text": "\n".join([f"{i+1}. {k}" for i, k in enumerate(kp_1_list)]) if isinstance(kp_1_list, list) else str(kp_1_list),
            "key_point_0_text": "\n".join([f"{i+1}. {k}" for i, k in enumerate(kp_0_list)]) if isinstance(kp_0_list, list) else str(kp_0_list),
            "answer": answer,
            "context": self._format_context(context) if context else ""
        })
        
        # 调用API
        response_text = self.api.call_llm(self.sp, up_content)
        
        # 解析JSON
        try:
            cleaned = clean_json_text(response_text)
            return json.loads(cleaned)
        except json.JSONDecodeError as e:
            raise Exception(f"评分响应解析失败: {str(e)}")


class QuestionClassifier:
    """问题类型分类器"""

    def __init__(self, api_client: APIClient, sp: str, up_template: str):
        self.api = api_client
        self.sp = sp
        self.up_template = up_template

    def classify(self, query: str, answer: str,
                 score_result: Dict[str, Any]) -> Dict[str, Any]:
        """根据评分结果对问题类型进行分类"""
        # 构建1分考点判定详情文本
        kpr_1_list = score_result.get("key_point_1_result", [])
        if isinstance(kpr_1_list, list):
            lines = []
            for kpr in kpr_1_list:
                point = kpr.get("point", "")
                satisfied = kpr.get("satisfied", "")
                evidence = kpr.get("evidence", "")
                lines.append(f"- 1分考点: {point}\n  是否满足: {satisfied}\n  判定依据: {evidence}")
            keypoint_1_results_text = "\n".join(lines) if lines else "无"
        else:
            keypoint_1_results_text = str(kpr_1_list)

        # 构建0分考点判定详情文本
        kpr_0_list = score_result.get("key_point_0_result", [])
        if isinstance(kpr_0_list, list):
            lines = []
            for kpr in kpr_0_list:
                point = kpr.get("point", "")
                satisfied = kpr.get("satisfied", "")
                evidence = kpr.get("evidence", "")
                lines.append(f"- 0分考点: {point}\n  是否满足: {satisfied}\n  判定依据: {evidence}")
            keypoint_0_results_text = "\n".join(lines) if lines else "无"
        else:
            keypoint_0_results_text = str(kpr_0_list)

        # 渲染用户提示词
        up_tpl = Template(self.up_template)
        up_content = up_tpl.render({
            "query": query,
            "answer": answer,
            "score": score_result.get("score", ""),
            "reason": score_result.get("reason", ""),
            "keypoint_1_results": keypoint_1_results_text,
            "keypoint_0_results": keypoint_0_results_text
        })

        # 调用API
        response_text = self.api.call_llm(self.sp, up_content)

        # 解析JSON
        try:
            cleaned = clean_json_text(response_text)
            return json.loads(cleaned)
        except json.JSONDecodeError as e:
            raise Exception(f"问题类型分类响应解析失败: {str(e)}")


class Evaluator:
    """评估器主类"""

    def __init__(self, config_file: str = "evaluator_config.json",
                 api_url: Optional[str] = None,
                 api_token: Optional[str] = None):
        # 加载配置
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        self.config = config
        
        # API配置
        self.api_url = api_url or config.get("api_url", "https://open.bigmodel.cn/api/paas/v4/chat/completions")
        self.api_token = api_token or config.get("api_token", "")
        
        # 创建API客户端
        api_client = APIClient(
            api_url=self.api_url,
            api_token=self.api_token,
            model_config=config.get("config", {})
        )
        
        # 创建考点生成器
        self.keypoint_generator = KeypointGenerator(
            api_client=api_client,
            sp=config.get("sp", ""),
            up_template=config.get("up", "")
        )
        
        # 创建评分器
        self.scorer = Scorer(
            api_client=api_client,
            sp=config.get("scorer_sp", ""),
            up_template=config.get("scorer_up", "")
        )
        
        # 创建问题类型分类器
        tagger_sp = config.get("tagger_sp", "")
        tagger_up = config.get("tagger_up", "")
        self.tagger = QuestionClassifier(
            api_client=api_client,
            sp=tagger_sp,
            up_template=tagger_up
        ) if tagger_sp and tagger_up else None

    def evaluate(self, query: str, answer: str,
                 date: Optional[str] = None,
                 demand: Optional[str] = None,
                 context: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
        """执行完整评估流程"""
        
        # Step 1: 生成考点
        keypoint_result = self.keypoint_generator.generate(
            query=query,
            date=date,
            demand=demand,
            context=context
        )
        
        # Step 2: 评分
        score_result = self.scorer.score(
            query=query,
            answer=answer,
            keypoint_result=keypoint_result,
            context=context
        )
        
        # Step 3: 问题类型分类（如果配置了分类器）
        tag_result = {}
        if self.tagger:
            try:
                tag_result = self.tagger.classify(
                    query=query,
                    answer=answer,
                    score_result=score_result
                )
            except Exception as e:
                tag_result = {"tag": f"分类失败: {str(e)}", "reason": ""}
        
        # 合并结果
        return {
            "keypoint_generation": keypoint_result,
            "scoring": score_result,
            "tagging": tag_result
        }


def load_excel_data(file_path: str) -> List[Dict[str, Any]]:
    """从Excel文件加载数据"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    headers = []
    data = []
    
    # 读取表头
    for cell in ws[1]:
        headers.append(cell.value)
    
    # 读取数据行
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        for idx, cell in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = cell.value
        data.append(row_data)
    
    return data


def flatten_result(result: Dict[str, Any]) -> Dict[str, Any]:
    """将嵌套的评估结果展开为平铺字典，所有字段独立成列"""
    flat = {}
    kp = result.get("keypoint_generation", {})
    sc = result.get("scoring", {})

    # --- 考点生成字段 ---
    flat["kp_thinking"] = kp.get("thinking", "")
    flat["kp_analysis"] = kp.get("analysis", "")
    flat["kp_main_demand"] = kp.get("main_demand", "")

    # key_point_1 数组（1分考点）
    kp_1_list = kp.get("key_point_1", [])
    if isinstance(kp_1_list, list):
        flat["kp_key_point_1"] = " | ".join(str(k) for k in kp_1_list) if kp_1_list else ""
    else:
        flat["kp_key_point_1"] = str(kp_1_list) if kp_1_list else ""

    # key_point_0 数组（0分考点）
    kp_0_list = kp.get("key_point_0", [])
    if isinstance(kp_0_list, list):
        flat["kp_key_point_0"] = " | ".join(str(k) for k in kp_0_list) if kp_0_list else ""
    else:
        flat["kp_key_point_0"] = str(kp_0_list) if kp_0_list else ""

    # --- 评分字段 ---
    flat["scorer_thinking"] = sc.get("thinking", "")
    flat["scorer_analysis"] = sc.get("analysis", "")
    flat["scorer_main_demand"] = sc.get("main_demand", "")

    # scorer_key_point：合并1分和0分考点列表
    sc_kp_1_list = sc.get("key_point_1", [])
    sc_kp_0_list = sc.get("key_point_0", [])
    scorer_key_point_lines = []
    # 1分条件
    if isinstance(sc_kp_1_list, list) and sc_kp_1_list:
        scorer_key_point_lines.append("【满足1分的条件】")
        for i, kp in enumerate(sc_kp_1_list, 1):
            scorer_key_point_lines.append(f"{i}. {kp}")
    elif sc_kp_1_list:
        scorer_key_point_lines.append("【满足1分的条件】")
        scorer_key_point_lines.append(f"1. {sc_kp_1_list}")
    # 0分条件
    if isinstance(sc_kp_0_list, list) and sc_kp_0_list:
        scorer_key_point_lines.append("\n【触发0分的条件】")
        for i, kp in enumerate(sc_kp_0_list, 1):
            scorer_key_point_lines.append(f"{i}. {kp}")
    elif sc_kp_0_list:
        scorer_key_point_lines.append("\n【触发0分的条件】")
        scorer_key_point_lines.append(f"1. {sc_kp_0_list}")
    flat["scorer_key_point"] = "\n".join(scorer_key_point_lines)

    # keypoint_results：合并1分和0分考点判定结果
    kpr_1_list = sc.get("key_point_1_result", [])
    kpr_0_list = sc.get("key_point_0_result", [])
    keypoint_results_lines = []
    # 1分考点判定结果
    if isinstance(kpr_1_list, list) and kpr_1_list:
        keypoint_results_lines.append("【1分考点判定结果】")
        for i, kpr in enumerate(kpr_1_list, 1):
            point = kpr.get("point", "")
            satisfied = kpr.get("satisfied", "")
            evidence = kpr.get("evidence", "")
            status = "✓ 满足" if str(satisfied).lower() in ("true", "1", "是", "满足") else "✗ 未满足"
            keypoint_results_lines.append(f"{i}. {point}")
            keypoint_results_lines.append(f"   状态: {status}")
            keypoint_results_lines.append(f"   依据: {evidence}")
            keypoint_results_lines.append("")
        # 统计
        total_1 = len(kpr_1_list)
        satisfied_1_count = sum(1 for k in kpr_1_list if str(k.get("satisfied", "")).lower() in ("true", "1", "是", "满足"))
        flat["keypoint_results_summary"] = f"1分考点: {satisfied_1_count}/{total_1}"
    elif kpr_1_list:
        keypoint_results_lines.append("【1分考点判定结果】")
        keypoint_results_lines.append(f"1. {json.dumps(kpr_1_list, ensure_ascii=False)}")
        flat["keypoint_results_summary"] = ""
    # 0分考点判定结果
    if isinstance(kpr_0_list, list) and kpr_0_list:
        keypoint_results_lines.append("\n【0分考点判定结果】")
        for i, kpr in enumerate(kpr_0_list, 1):
            point = kpr.get("point", "")
            satisfied = kpr.get("satisfied", "")
            evidence = kpr.get("evidence", "")
            status = "✓ 触发" if str(satisfied).lower() in ("true", "1", "是", "满足", "触发") else "✗ 未触发"
            keypoint_results_lines.append(f"{i}. {point}")
            keypoint_results_lines.append(f"   状态: {status}")
            keypoint_results_lines.append(f"   依据: {evidence}")
            keypoint_results_lines.append("")
        # 统计
        total_0 = len(kpr_0_list)
        triggered_0_count = sum(1 for k in kpr_0_list if str(k.get("satisfied", "")).lower() in ("true", "1", "是", "满足", "触发"))
        if flat.get("keypoint_results_summary"):
            flat["keypoint_results_summary"] += f", 0分考点触发: {triggered_0_count}/{total_0}"
        else:
            flat["keypoint_results_summary"] = f"0分考点触发: {triggered_0_count}/{total_0}"
    elif kpr_0_list:
        keypoint_results_lines.append("\n【0分考点判定结果】")
        keypoint_results_lines.append(f"1. {json.dumps(kpr_0_list, ensure_ascii=False)}")
    flat["keypoint_results"] = "\n".join(keypoint_results_lines)

    # 评分核心字段
    flat["score"] = sc.get("score", "")
    flat["reason"] = sc.get("reason", "")

    # --- 问题类型分类字段 ---
    tg = result.get("tagging", {})
    flat["question_type"] = tg.get("tag", "")
    flat["question_type_reason"] = tg.get("reason", "")

    return flat


def save_excel_data(data: List[Dict[str, Any]], output_path: str):
    """保存数据到Excel文件"""
    if not data:
        print("警告: 没有数据需要保存")
        return

    wb = openpyxl.Workbook()
    ws = wb.active

    # 收集所有出现过的字段名（保持顺序）
    headers = list(data[0].keys())
    for item in data[1:]:
        for key in item.keys():
            if key not in headers:
                headers.append(key)

    # 写入表头
    ws.append(headers)

    # 写入数据
    for item in data:
        row = []
        for key in headers:
            value = item.get(key, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            elif value is None:
                value = ""
            row.append(value)
        ws.append(row)

    wb.save(output_path)
    print(f"结果已保存到: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='智谱AI评估器')
    parser.add_argument('--query', type=str, help='问题')
    parser.add_argument('--answer', type=str, help='回答')
    parser.add_argument('--date', type=str, help='日期信息')
    parser.add_argument('--demand', type=str, help='需求信息')
    parser.add_argument('--excel', type=str, help='批量评估Excel文件路径')
    parser.add_argument('--output', type=str, help='输出文件路径')
    parser.add_argument('--api_key', type=str, help='豆包API Key')
    parser.add_argument('--api_url', type=str, help='豆包API URL')
    parser.add_argument('--config', type=str, default='豆包评估配置.json',
                       help='配置文件路径')
    
    args = parser.parse_args()
    
    # 创建评估器
    evaluator = Evaluator(
        config_file=args.config,
        api_url=args.api_url,
        api_token=args.api_key or "e6fc146d-4b6f-4803-bf53-a530b354bd23"
    )
    
    # 单条评估模式
    if args.query and args.answer:
        print(f"\n{'='*60}")
        print(f"评估问题: {args.query}")
        print(f"{'='*60}\n")
        
        result = evaluator.evaluate(
            query=args.query,
            answer=args.answer,
            date=args.date,
            demand=args.demand
        )
        
        print("\n✅ 评估结果:")
        print(json.dumps(result, indent=2, ensure_ascii=False))
        
        # 如果需要保存结果
        if args.output:
            flat = flatten_result(result)
            output_data = [{
                "query": args.query,
                "answer": args.answer,
                **flat,
                "evaluated_at": datetime.now().isoformat()
            }]
            save_excel_data(output_data, args.output)
    
    # 批量评估模式
    elif args.excel:
        print(f"\n{'='*60}")
        print(f"批量评估模式: {args.excel}")
        print(f"{'='*60}\n")
        
        # 加载数据
        data = load_excel_data(args.excel)
        print(f"加载了 {len(data)} 条数据\n")
        
        results = []
        session_history = {}
        
        for idx, item in enumerate(data, 1):
            query = item.get('query', item.get('问题', ''))
            answer = item.get('answer', item.get('回答', ''))
            session_id = item.get('session_id', item.get('session', None))
            
            if not query or not answer:
                print(f"⚠️  第{idx}条: 缺少query或answer，跳过")
                continue
            
            print(f"[{idx}/{len(data)}] 评估: {query[:50]}...")
            
            try:
                # 获取上下文
                context = session_history.get(session_id, []) if session_id else None
                
                # 评估
                result = evaluator.evaluate(
                    query=query,
                    answer=answer,
                    date=item.get('date', None),
                    demand=item.get('demand', None),
                    context=context
                )
                
                # 保存到历史
                if session_id:
                    if session_id not in session_history:
                        session_history[session_id] = []
                    session_history[session_id].append({
                        'query': query,
                        'answer': answer,
                        'keypoints': result['keypoint_generation']
                    })
                
                # 汇总结果 - 展开为独立列
                flat = flatten_result(result)
                result_item = {
                    **item,
                    **flat,
                    'evaluated_at': datetime.now().isoformat()
                }
                results.append(result_item)
                
                print(f"✅ 完成\n")
                
                # 延迟避免速率限制
                time.sleep(0.5)
                
            except Exception as e:
                print(f"❌ 失败: {str(e)}\n")
                results.append({
                    **item,
                    'error': str(e),
                    'evaluated_at': datetime.now().isoformat()
                })
        
        # 保存结果
        output_path = args.output or f"eval_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_excel_data(results, output_path)
        
        print(f"\n{'='*60}")
        print(f"批量评估完成!")
        print(f"成功: {len([r for r in results if 'error' not in r])} 条")
        print(f"失败: {len([r for r in results if 'error' in r])} 条")
        print(f"结果保存到: {output_path}")
        print(f"{'='*60}\n")
    
    else:
        parser.print_help()
        print("\n使用示例:")
        print("  # 单条评估")
        print("  python standalone_evaluator_zhipu.py --query '什么是AI?' --answer 'AI是...'")
        print("\n  # 批量评估")
        print("  python standalone_evaluator_zhipu.py --excel input.xlsx --output result.xlsx")


if __name__ == "__main__":
    main()
